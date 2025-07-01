import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import Font

def export_forms_to_excel():
    os.makedirs("reports", exist_ok=True)
    path = os.path.join("reports", "forms.xlsx")

    conn = sqlite3.connect("db/forms.db")
    cursor = conn.cursor()
    cursor. execute("SELECT * FROM forms")
    rows = cursor.fetchall()
    conn.close()

    headers = ["ID", "Name", "Email", "Phone", "Message", "City"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Анкеты"

    # Заголовки с жирным шрифтом
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, colums=col_num, value=header)
        cell.font = Font(bold=True)

    # Данные
    for row_idX, row in enumerate(rows, start=2):
        for col_idX, value in enumerate(row, start=1):
            ws.cell(row=row_idX, column=col_idX, value=value)

    wb.save(path)
    print(f"✅ Excel-отчёт успешно создан: {path}")